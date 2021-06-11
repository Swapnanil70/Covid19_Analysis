import  openpyxl as xl
import statistics as st
import math
import scipy.stats as ss
import  numpy as np

wb = xl.load_workbook('covid19-dar.xlsx')
sheet = wb['COVID-19-geographic-disbtributi']

cases_china = []
cases_italy = []
sam_cases_china = []
sam_cases_italy = []

for row in range(2, 194):
    cell = sheet.cell(row, 12)
    country = cell.value
    if country == 'China':
        cell = sheet.cell(row, 13)
        tot_cases_china = cell.value

    if country == 'Italy':
        cell = sheet.cell(row, 13)
        tot_cases_italy = cell.value

pop_mean_china = tot_cases_china/87
pop_mean_italy = tot_cases_italy/56

for row in range(1331, 1417):
    cell = sheet.cell(row, 5)
    cases_china.append(cell.value)

for row in range(3371,3427):
    cell = sheet.cell(row, 5)
    cases_italy.append(cell.value)

pop_std_china = st.stdev(cases_china)
pop_std_italy = st.stdev(cases_italy)

for row in range(1331, 1395):
    cell = sheet.cell(row, 5)
    sam_cases_china.append(cell.value)

for row in range(3371, 3389):
    cell = sheet.cell(row, 5)
    sam_cases_italy.append(cell.value)

sam_mean_china = sum(sam_cases_china)/len(sam_cases_china)
sam_mean_italy = sum(sam_cases_italy)/len(sam_cases_italy)

sam_std_china = st.stdev(sam_cases_china)
sam_std_italy = st.stdev(sam_cases_italy)

t_statistic_china = (sam_mean_china - pop_mean_china)/(sam_std_china/math.sqrt(len(sam_cases_china)))
t_statistic_italy = (sam_mean_italy - pop_mean_italy)/(sam_std_italy/math.sqrt(len(sam_cases_italy)))
dof_china = len(sam_cases_china)-1
dof_italy = len(sam_cases_italy)-1

p_value_china = ss.t.sf(np.abs(t_statistic_china), dof_china)
p_value_italy = ss.t.sf(np.abs(t_statistic_italy), dof_italy)

print('Sample Mean China:- ',sam_mean_china,' Null Hypothesis:- ', pop_mean_china,' Sample S.D:- ',sam_std_china)
print('Sample Mean Italy:- ',sam_mean_italy,' Null Hypothesis:- ', pop_mean_italy,' Sample S.D:- ',sam_std_italy)

print(' T- statistic for China:- ', t_statistic_china,' corresponding p value:- ', p_value_china,' with degrees of freedom = ', dof_china)
print(' T- statistic for Italy:- ', t_statistic_italy,' corresponding p value:- ', p_value_italy,' with degrees of freedom = ', dof_italy)

if p_value_china > 0.05:
    print('Null Hypothesis Accepted, spread is reduced due to lockdown in China')

else:
    print('Null Hypothesis is rejected, spread increased in China')

if p_value_italy > 0.05:
    print('Null Hypothesis Accepted, spread is reduced due to lockdown in Italy')

else:
    print('Null Hypothesis is rejected, spread increased in Italy')

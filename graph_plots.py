import openpyxl as xl
import matplotlib.pyplot as plt

wb = xl.load_workbook('covid19-dar.xlsx')
sheet = wb['COVID-19-geographic-disbtributi']

def fatality_rate():

    f_rate = []
    countries = []

    for row in range(2,194): #Extracting values from the given data base
        cell = sheet.cell(row, 15)
        if cell.value > 0:
            countries.append(cell.value)
            cell = sheet.cell(row, 12)
            f_rate.append(cell.value)

    plt.bar(f_rate, countries, width = 0.8, color = ['blue'] )
    plt.xticks(rotation = 90, fontsize = 5)

    plt.xlabel('Countries')
    plt.ylabel('Fatality Rate')

    plt.title('Bar chart representing fatality rate in various countries due to COVID-19')
    plt.show()

def total_cases():

    cases = []
    countries = []

    for row in range (2,194): #Extracting values from the given data base
        cell = sheet.cell(row, 12)
        countries.append(cell.value)
        cell = sheet.cell(row, 13)
        cases.append(cell.value)

    n = len(cases)

    for i in range(n): #Bubble Sort Algorithm
        for j in range(0, n-i-1):
            if cases[j] < cases[j+1]:
                cases[j], cases[j+1] = cases[j+1], cases[j]
                countries[j], countries[j+1] = countries[j+1], countries[j]

    for i in range(len(countries)):
        if countries[i] == 'India':
            pos = i
            break

    plt.bar(countries, cases, width = 0.8, color = ['blue'])
    plt.xticks(rotation = 90, fontsize = 5)

    plt.xlabel('Countries')
    plt.ylabel('Cases')

    plt.title('Bar chart representing the increase in total cases and the position of India :-')
    print('The position of India based on total number of cases is :- ', pos)
    plt.show()

def mean_daily_cases():

    mean_d_cases = []
    countries = []

    for row in range(2, 194): #Extracting the data from database
        cell = sheet.cell(row, 12)
        countries.append(cell.value)
        cell = sheet.cell(row , 13)
        mean_d_cases.append(cell.value/86)
        cell = sheet.cell(row, 16)
        cell.value = mean_d_cases[row-2]

    n = len(mean_d_cases)

    for i in range(n):  # Bubble Sort Algorithm
        for j in range(0, n - i - 1):
            if mean_d_cases[j] < mean_d_cases[j + 1]:
                mean_d_cases[j], mean_d_cases[j + 1] = mean_d_cases[j + 1], mean_d_cases[j]
                countries[j], countries[j + 1] = countries[j + 1], countries[j]

    for i in range(len(countries)):
        if countries[i] == 'India':
            pos = i
            break

    plt.bar(countries, mean_d_cases, width=0.7, color=['blue'])
    plt.xticks(rotation=90, fontsize=5)

    plt.xlabel('Countries')
    plt.ylabel('Mean Daily Cases')

    plt.title('Bar chart representing the increase in mean daily cases and the position of India :-')
    print('The position of India based on mean daily cases is :- ', pos)
    plt.show()
    wb.save('covid19-dar.xlsx')

total_cases()
fatality_rate()
mean_daily_cases()

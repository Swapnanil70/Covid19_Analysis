import openpyxl as xl

wb = xl.load_workbook('covid19-dar.xlsx')
sheet = wb['COVID-19-geographic-disbtributi']

i,sum_c,sum_d,no_of_countries,max_f_rate = 0,0,0,0,0
countries = []
cases = []
deaths = []
max_f_rate_country = ''

for row in range(2, sheet.max_row +1):
    cell = sheet.cell(row, 7)
    countries.append(cell.value)
    cell2 = sheet.cell(row, 5)
    cases.append(cell2.value)
    cell3 = sheet.cell(row,6)
    deaths.append(cell3.value)

for c in range(len(countries)-2):
    if countries[c] != countries[c+1]:
        no_of_countries+=1

for row in range(2, no_of_countries+2):
    while  countries[i] == countries[i+1] and i<len(countries):
        sum_c += cases[i]
        sum_d += deaths[i]
        i+=1

    countries_c = sheet.cell(row,12)
    cases_c = sheet.cell(row,13)
    deaths_d = sheet.cell(row,14)
    countries_c.value = countries[i-1]
    cases_c.value = sum_c
    deaths_d.value = sum_d
    fatality_rate = sheet.cell(row, 15)
    if sum_c > 0:
        f_rate = (float)(sum_d/sum_c)*100
        fatality_rate.value = f_rate
        if f_rate > max_f_rate and f_rate < 15:
            max_f_rate = f_rate
            max_f_rate_country = countries[i-1]
    else:
        fatality_rate.value = 0
    i+=1
    sum_d = 0
    sum_c = 0

print("The country with the maximum fatality rate is:- ",max_f_rate_country," with a fatality rate of:- ",max_f_rate)

wb.save('covid19-dar.xlsx')





import openpyxl as xl
import statistics as st
import math
import scipy.stats as ss

wb = xl.load_workbook('covid19-dar.xlsx')
sheet = wb['COVID-19-geographic-disbtributi']

cases_India = []
sam_cases = []

for row in range(2, 194):
    cell = sheet.cell(row, 12)
    country = cell.value
    if country == 'India':
        cell = sheet.cell(row, 13)
        tot_cases = cell.value

pop_mean = tot_cases/57

for row in range(2858, 2914):
    cell = sheet.cell(row, 5)
    cases_India.append(cell.value)

pop_std = st.stdev(cases_India)

for row in range(2858, 2863):
    cell = sheet.cell(row, 5)
    sam_cases.append(cell.value)

sam_mean = sum(sam_cases)/len(sam_cases)
sam_std = st.stdev(sam_cases)

z_score = (sam_mean-pop_mean)/(pop_std/math.sqrt(5))
p_value = ss.norm.ppf(z_score)

print('Population Mean:- ',pop_mean,'\nPopulation Standard Deviation:- ',pop_std,'\nSample Mean:- ',sam_mean)
print('The z-score is :- ', z_score)
print('P value for z-test:-')
print('P value is:- ',p_value)

if p_value > 0.05:
    print('Null Hypothesis of decrease in spread is satisfied in India')

else:
    print('Null Hypothesis is rejected')